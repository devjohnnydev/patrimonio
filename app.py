from flask import Flask, render_template, redirect, url_for, request, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from models import db, Escola, User, Sala, Patrimonio, Inventario, ItemInventario, SolicitacaoRealocacao, MensagemChat
from datetime import datetime
import os
import openpyxl
from sqlalchemy import func

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-12345')

# Configuração para Railway / PostgreSQL
database_url = os.environ.get('DATABASE_URL')
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url or 'sqlite:///patrimonio.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- Configuração de Email (Flask-Mail via SMTP) ---
app.config['MAIL_SERVER']   = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT']     = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']  = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@senai.br')

mail = Mail(app)
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.context_processor
def inject_notifications():
    if current_user.is_authenticated:
        q = MensagemChat.query.filter_by(escola_id=current_user.escola_id, lida=False).filter(MensagemChat.usuario_id != current_user.id)
        if current_user.role == 'professor':
            sala_ids = [s.id for s in current_user.salas]
            unread_count = q.filter(MensagemChat.sala_id.in_(sala_ids)).count()
        else:
            unread_count = q.count()
        return dict(unread_count=unread_count)
    return dict(unread_count=0)

# --- Auxiliares de E-mail ---
def enviar_email(destinatario, assunto, corpo_html):
    """Envia e-mail real via Flask-Mail. Faz log se falhar."""
    try:
        msg = Message(assunto, recipients=[destinatario], html=corpo_html)
        mail.send(msg)
        print(f"[EMAIL ENVIADO] Para: {destinatario} | Assunto: {assunto}")
    except Exception as e:
        print(f"[EMAIL ERRO] Falha ao enviar para {destinatario}: {e}")

def corpo_boas_vindas(user, senha_pura, escola_nome, base_url):
    salas = ', '.join([s.nome for s in user.salas]) or 'Nenhum ambiente designado ainda'
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;">
      <div style="background:#cc0000;padding:24px;text-align:center;">
        <h1 style="color:white;margin:0;font-size:28px;">SENAI</h1>
        <p style="color:rgba(255,255,255,0.8);margin:4px 0 0;">Gestão de Patrimônio</p>
      </div>
      <div style="padding:32px;background:#f9f9f9;">
        <h2 style="color:#222;">Olá, {user.nome}! 👋</h2>
        <p style="color:#555;">Seu acesso à plataforma de <strong>Gestão de Patrimônio SENAI</strong> foi criado com sucesso.</p>
        <div style="background:#fff;border:1px solid #eee;border-radius:8px;padding:20px;margin:20px 0;">
          <p style="margin:0 0 8px;"><strong>🏫 Unidade:</strong> {escola_nome}</p>
          <p style="margin:0 0 8px;"><strong>👤 Usuário:</strong> <code style="background:#f0f0f0;padding:2px 6px;border-radius:4px;">{user.username}</code></p>
          <p style="margin:0 0 8px;"><strong>🔑 Senha:</strong> <code style="background:#f0f0f0;padding:2px 6px;border-radius:4px;">{senha_pura}</code></p>
          <p style="margin:0;"><strong>🏠 Ambientes:</strong> {salas}</p>
        </div>
        <div style="text-align:center;margin:24px 0;">
          <a href="{base_url}" style="background:#cc0000;color:white;padding:14px 32px;border-radius:6px;text-decoration:none;font-weight:bold;">ACESSAR PLATAFORMA</a>
        </div>
        <p style="color:#999;font-size:12px;">Este é um e-mail automático. Não responda a esta mensagem.</p>
      </div>
    </div>
    """

def corpo_reset_senha(user, link):
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;">
      <div style="background:#cc0000;padding:24px;text-align:center;">
        <h1 style="color:white;margin:0;font-size:28px;">SENAI</h1>
        <p style="color:rgba(255,255,255,0.8);margin:4px 0 0;">Gestão de Patrimônio</p>
      </div>
      <div style="padding:32px;background:#f9f9f9;">
        <h2 style="color:#222;">Redefinição de Senha</h2>
        <p style="color:#555;">Olá, <strong>{user.nome}</strong>. Uma solicitação de redefinição de senha foi feita para sua conta.</p>
        <div style="text-align:center;margin:32px 0;">
          <a href="{link}" style="background:#cc0000;color:white;padding:14px 32px;border-radius:6px;text-decoration:none;font-weight:bold;">REDEFINIR MINHA SENHA</a>
        </div>
        <p style="color:#999;font-size:12px;">Este link expira em <strong>1 hora</strong>. Se não foi você, ignore este e-mail.</p>
        <p style="color:#bbb;font-size:11px;">SENAI São Paulo — Gestão de Patrimônio SaaS</p>
      </div>
    </div>
    """

@app.template_filter('status_badge')
def status_badge(status):
    badges = {
        'pendente': 'secondary',
        'aprovado': 'success',
        'recusado': 'danger',
        'ativo': 'success',
        'inativo': 'warning',
        'confirmado': 'success',
        'alterado': 'info',
        'fora_de_lugar': 'warning',
        'nao_encontrado': 'danger',
        'concluido': 'primary',
        'validado': 'dark'
    }
    return badges.get(status, 'primary')

import os
from werkzeug.utils import secure_filename

# --- Perfil e Identidade ---
@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    if request.method == 'POST':
        nova_senha = request.form.get('password')
        
        if nova_senha:
            current_user.set_password(nova_senha)
            
        # Upload de Foto
        if 'foto_file' in request.files:
            file = request.files['foto_file']
            if file and file.filename != '':
                filename = secure_filename(f"user_{current_user.id}_{file.filename}")
                upload_dir = os.path.join(app.root_path, 'static/uploads/perfil')
                os.makedirs(upload_dir, exist_ok=True)
                
                full_path = os.path.join(upload_dir, filename)
                file.save(full_path)
                current_user.foto_url = f"/static/uploads/perfil/{filename}"
            
        db.session.commit()
        flash('Perfil atualizado com sucesso!', 'success')
        return redirect(url_for('perfil'))
        
    return render_template('perfil.html')

@app.route('/notificacoes')
@login_required
def notificacoes():
    q = MensagemChat.query.filter_by(escola_id=current_user.escola_id, lida=False).filter(MensagemChat.usuario_id != current_user.id)
    if current_user.role == 'professor':
        sala_ids = [s.id for s in current_user.salas]
        unread_msgs = q.filter(MensagemChat.sala_id.in_(sala_ids)).all()
    else:
        unread_msgs = q.all()
    
    alertas = {}
    for msg in unread_msgs:
        if msg.sala_id not in alertas:
            alertas[msg.sala_id] = {'sala': msg.sala, 'count': 0, 'data': msg.data_hora}
        alertas[msg.sala_id]['count'] += 1
    return render_template('notificacoes.html', alertas=alertas.values())

def get_sala_accuracy(sala_id):
    total = Patrimonio.query.filter_by(sala_id=sala_id, status='ativo').count()
    if total == 0: return 100
    inv = Inventario.query.filter_by(sala_id=sala_id).order_by(Inventario.data_hora_inicio.desc()).first()
    if not inv: return 0
    conferidos = ItemInventario.query.filter_by(inventario_id=inv.id).filter(ItemInventario.status.in_(['confirmado', 'alterado'])).count()
    return min(100, round((conferidos / total) * 100))

# --- Rotas de Autenticação ---
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role in ['admin', 'coordenador']:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('professor_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('index'))
        flash('Usuário ou senha inválidos', 'danger')
    return render_template('login.html')

# --- Esqueci minha senha ---
@app.route('/esqueci-senha', methods=['GET', 'POST'])
def esqueci_senha():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user:
            token = serializer.dumps(user.email, salt='reset-senha')
            link = url_for('resetar_senha', token=token, _external=True)
            enviar_email(user.email, 'Redefinição de Senha — SENAI Patrimônio', corpo_reset_senha(user, link))
        # Sempre mostra a mesma mensagem por segurança
        flash('Se o e-mail estiver cadastrado, você receberá as instruções em breve.', 'info')
        return redirect(url_for('login'))
    return render_template('esqueci_senha.html')

@app.route('/resetar-senha/<token>', methods=['GET', 'POST'])
def resetar_senha(token):
    try:
        email = serializer.loads(token, salt='reset-senha', max_age=3600)
    except (SignatureExpired, BadSignature):
        flash('Link inválido ou expirado. Solicite um novo.', 'danger')
        return redirect(url_for('esqueci_senha'))
    
    user = User.query.filter_by(email=email).first_or_404()
    
    if request.method == 'POST':
        nova = request.form.get('password')
        confirma = request.form.get('confirm_password')
        if nova != confirma:
            flash('As senhas não coincidem.', 'danger')
        elif len(nova) < 6:
            flash('A senha deve ter pelo menos 6 caracteres.', 'danger')
        else:
            user.set_password(nova)
            db.session.commit()
            flash('Senha alterada com sucesso! Faça login.', 'success')
            return redirect(url_for('login'))
    return render_template('resetar_senha.html', token=token)

# --- Admin: Enviar link de reset para qualquer usuário ---
@app.route('/admin/resetar-senha/<int:user_id>')
@login_required
def admin_resetar_senha(user_id):
    if current_user.role not in ['admin', 'coordenador']:
        return redirect(url_for('index'))
    user = User.query.filter_by(id=user_id, escola_id=current_user.escola_id).first_or_404()
    token = serializer.dumps(user.email, salt='reset-senha')
    link = url_for('resetar_senha', token=token, _external=True)
    enviar_email(user.email, 'Redefinição de Senha — SENAI Patrimônio', corpo_reset_senha(user, link))
    flash(f'Link de redefinição enviado para {user.email}!', 'success')
    return redirect(url_for('admin_responsaveis'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- Dashboard Admin ---
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role not in ['admin', 'coordenador']:
        return redirect(url_for('index'))
    
    e_id = current_user.escola_id
    
    # Métricas Avançadas
    stats = {
        'salas': Sala.query.filter_by(escola_id=e_id).count(),
        'patrimonios': Patrimonio.query.filter_by(escola_id=e_id).count(),
        'quebrados': Patrimonio.query.filter_by(escola_id=e_id, status_conservacao='quebrado').count(),
        'relocacoes_pendentes': SolicitacaoRealocacao.query.filter_by(escola_id=e_id, status='pendente').count()
    }
    
    # Quantidade por Sala
    itens_por_sala = db.session.query(Sala.nome, func.count(Patrimonio.id)).\
        join(Patrimonio, Patrimonio.sala_id == Sala.id).\
        filter(Sala.escola_id == e_id).\
        group_by(Sala.nome).all()

    # Status de Conservação
    conservacao_stats = db.session.query(Patrimonio.status_conservacao, func.count(Patrimonio.id)).\
        filter(Patrimonio.escola_id == e_id).\
        group_by(Patrimonio.status_conservacao).all()

    # Produtividade por Responsável
    produtividade = db.session.query(User.nome, func.count(Inventario.id)).\
        join(Inventario, Inventario.responsavel_id == User.id).\
        filter(User.escola_id == e_id).\
        group_by(User.nome).all()

    salas = Sala.query.filter_by(escola_id=e_id).all()
    relocacoes = SolicitacaoRealocacao.query.filter_by(escola_id=e_id, status='pendente').all()
    quebrados = Patrimonio.query.filter_by(escola_id=e_id, status_conservacao='quebrado').all()
    inventarios_concluidos = Inventario.query.filter_by(escola_id=e_id, status='concluido').all()
    
    salas_data = []
    total_ativos_escola = 0
    total_encontrados_escola = 0
    
    for s in salas:
        # Itens ativos na sala
        ativos_sala = Patrimonio.query.filter_by(sala_id=s.id, status='ativo').count()
        total_ativos_escola += ativos_sala
        
        # Acurácia
        acuracia = get_sala_accuracy(s.id)
        if ativos_sala > 0:
            # Encontrados = acuracia * ativos / 100
            total_encontrados_escola += int((acuracia / 100) * ativos_sala)
            
        salas_data.append({
            'obj': s,
            'acuracia': acuracia
        })

    # Ranking de Intervenção (Piores 5)
    top_criticos = sorted(salas_data, key=lambda x: x['acuracia'])[:5]
    
    # Eficiência Global
    global_accuracy = min(100, round((total_encontrados_escola / total_ativos_escola) * 100)) if total_ativos_escola > 0 else 100
    
    # Itens Sumidos (Exemplo: Itens pendentes em balanços ativos)
    alertas_perda = db.session.query(Patrimonio).\
        join(Sala, Patrimonio.sala_id == Sala.id).\
        filter(Sala.escola_id == e_id, Patrimonio.status == 'ativo').\
        limit(10).all() 
    # Em um cenário real, cruzaríamos com os ItemInventario. Aqui pegamos ativos aleatórios para o mockup visual se necessário.

    return render_template('admin/dashboard.html', 
                           stats=stats, 
                           salas_meta=salas_data, 
                           top_criticos=top_criticos,
                           global_accuracy=global_accuracy,
                           alertas_perda=alertas_perda,
                           relocacoes=relocacoes, 
                           quebrados=quebrados,
                           itens_por_sala=itens_por_sala,
                           conservacao_stats=conservacao_stats,
                           produtividade=produtividade,
                           inventarios_concluidos=inventarios_concluidos)

@app.route('/admin/sala/<int:sala_id>')
@login_required
def admin_detalhes_sala(sala_id):
    if current_user.role not in ['admin', 'coordenador']:
        return redirect(url_for('index'))
    
    sala = Sala.query.filter_by(id=sala_id, escola_id=current_user.escola_id).first_or_404()
    
    # Itens que DEVERIAM estar aqui
    patrimonios = Patrimonio.query.filter_by(sala_id=sala_id, status='ativo').all()
    
    # Último inventário para pegar o status real
    inv = Inventario.query.filter_by(sala_id=sala_id).order_by(Inventario.data_hora_inicio.desc()).first()
    
    # Mapear status dos itens
    audit_data = []
    conferidos_count = 0
    
    # Itens da sala
    for p in patrimonios:
        item_inv = None
        if inv:
            item_inv = ItemInventario.query.filter_by(inventario_id=inv.id, patrimonio_id=p.id).first()
        
        status = 'pendente'
        data_validacao = None
        if item_inv:
            status = item_inv.status
            conferidos_count += 1
            # Para simplificar, pegamos a data do inventário se não houver no item
            data_validacao = inv.data_hora_fim or inv.data_hora_inicio 

        audit_data.append({
            'patrimonio': p,
            'status': status,
            'data_validacao': data_validacao
        })
    
    # Itens de FORA encontrados aqui (Conflitos)
    conflitos = []
    if inv:
        conflitos_inv = ItemInventario.query.filter_by(inventario_id=inv.id).all()
        for ci in conflitos_inv:
            if ci.patrimonio and ci.patrimonio.sala_id != sala_id:
                conflitos.append(ci)

    acuracia = min(100, round((conferidos_count / len(patrimonios)) * 100)) if patrimonios else 100
    
    return render_template('admin/detalhes_sala.html', 
                           sala=sala, 
                           audit_data=audit_data,
                           conflitos=conflitos,
                           acuracia=acuracia,
                           inventario=inv)

@app.route('/sala/exportar/<int:sala_id>')
@login_required
def exportar_sala_excel(sala_id):
    sala = Sala.query.get_or_404(sala_id)
    # Segurança: Verificar se pertence à escola
    if sala.escola_id != current_user.escola_id:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('index'))
    
    # Segurança: Se for responsável, verificar se a sala é dele
    if current_user.role == 'responsavel' and sala not in current_user.salas:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('index'))

    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Patrimonios"
    
    headers = ["Patrimonio", "Descricao", "Marca", "Modelo", "Status Conservacao"]
    sheet.append(headers)
    
    patrimonios = Patrimonio.query.filter_by(sala_id=sala_id, status='ativo').all()
    for p in patrimonios:
        sheet.append([p.numero_patrimonio, p.descricao, p.marca, p.modelo, p.status_conservacao])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"patrimonio_sala_{sala.nome.replace(' ', '_')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/importar', methods=['POST'])
@login_required
def admin_importar_excel():
    if current_user.role != 'admin': return redirect(url_for('index'))
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado', 'danger')
        return redirect(url_for('admin_patrimonios'))
    
    file = request.files['file']
    sala_id = request.form.get('sala_id')
    
    if not sala_id:
        flash('Selecione uma sala de destino.', 'danger')
        return redirect(url_for('admin_patrimonios'))
        
    sala_destino = Sala.query.filter_by(id=sala_id, escola_id=current_user.escola_id).first_or_404()

    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        e_id = current_user.escola_id
        count = 0
        
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        def get_val(row, col_name):
            try:
                idx = headers.index(col_name)
                return str(row[idx].value).strip() if row[idx].value else ""
            except:
                return ""

        for row in sheet.iter_rows(min_row=2):
            if not row[0].value: continue 
            
            pat_num = get_val(row, "Patrimonio") or str(row[0].value)
            desc = get_val(row, "Descricao") or str(row[1].value)
            marca = get_val(row, "Marca") or (str(row[2].value) if len(row) > 2 else "")
            modelo = get_val(row, "Modelo") or (str(row[3].value) if len(row) > 3 else "")
            # sala_nome = get_val(row, "Sala") # IGNORADO EM FAVOR DA SELEÇÃO DO UI
            foto_url = get_val(row, "Foto") or (str(row[5].value) if len(row) > 5 else "")
            
            # Evitar duplicados por número
            existente = Patrimonio.query.filter_by(numero_patrimonio=pat_num, escola_id=e_id).first()
            if not existente:
                novo_pat = Patrimonio(
                    numero_patrimonio=pat_num,
                    descricao=desc,
                    marca=marca,
                    modelo=modelo,
                    imagem_url=foto_url,
                    sala_id=sala_destino.id,
                    escola_id=e_id
                )
                db.session.add(novo_pat)
                count += 1
        
        db.session.commit()
        flash(f'Sucesso! {count} itens importados via Excel.', 'success')
    except Exception as e:
        flash(f'Erro ao processar Excel: {e}', 'danger')
        
    return redirect(url_for('admin_patrimonios'))

# CRUD Salas
@app.route('/admin/salas', methods=['GET', 'POST'])
@login_required
def admin_salas():
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    e_id = current_user.escola_id
    if request.method == 'POST':
        nova_sala = Sala(
            nome=request.form.get('nome'),
            bloco=request.form.get('bloco'),
            descricao=request.form.get('descricao'),
            imagem_url=request.form.get('imagem_url'),
            escola_id=e_id
        )
        # Vincular Professor(es) se selecionado
        prof_ids = request.form.getlist('professor_ids')
        for pid in prof_ids:
            u = User.query.filter_by(id=pid, escola_id=e_id).first()
            if u: nova_sala.responsaveis.append(u)
            
        db.session.add(nova_sala)
        db.session.commit()
        flash('Ambiente cadastrado e professor vinculado!', 'success')
        
    salas = Sala.query.filter_by(escola_id=e_id).all()
    professores = User.query.filter_by(role='professor', escola_id=e_id).all()
    return render_template('admin/salas.html', salas=salas, professores=professores)

@app.route('/admin/salas/editar/<int:id>', methods=['POST'])
@login_required
def admin_editar_sala(id):
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    sala = Sala.query.filter_by(id=id, escola_id=current_user.escola_id).first_or_404()
    
    sala.nome = request.form.get('nome')
    sala.bloco = request.form.get('bloco')
    sala.descricao = request.form.get('descricao')
    sala.imagem_url = request.form.get('imagem_url')
    
    # Atualizar Professor Responsável
    sala.responsaveis = []
    prof_ids = request.form.getlist('professor_ids')
    for pid in prof_ids:
        u = User.query.filter_by(id=pid, escola_id=current_user.escola_id).first()
        if u: sala.responsaveis.append(u)
        
    db.session.commit()
    flash('Ambiente atualizado!', 'success')
    return redirect(url_for('admin_salas'))

# CRUD Patrimonio
@app.route('/admin/patrimonios', methods=['GET', 'POST'])
@login_required
def admin_patrimonios():
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    e_id = current_user.escola_id
    if request.method == 'POST':
        novo_pat = Patrimonio(
            numero_patrimonio=request.form.get('numero'),
            descricao=request.form.get('descricao'),
            marca=request.form.get('marca'),
            modelo=request.form.get('modelo'),
            sala_id=request.form.get('sala_id'),
            imagem_url=request.form.get('imagem_url'),
            escola_id=e_id
        )
        db.session.add(novo_pat)
        db.session.commit()
    patrimonios = Patrimonio.query.filter_by(escola_id=e_id).all()
    salas = Sala.query.filter_by(escola_id=e_id).all()
    return render_template('admin/patrimonios.html', patrimonios=patrimonios, salas=salas)

# Fluxo de Descarte
@app.route('/admin/descartar/<int:id>')
@login_required
def admin_descartar(id):
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    pat = Patrimonio.query.filter_by(id=id, escola_id=current_user.escola_id).first_or_404()
    pat.status = 'descartado'
    pat.status_conservacao = 'descartado'
    db.session.commit()
    flash('Item enviado para descarte!', 'success')
    return redirect(url_for('admin_dashboard'))

# Balanço de Inventário (Resultados Consolidados)
@app.route('/admin/balanco/<int:inv_id>')
@login_required
def admin_balanco_inventario(inv_id):
    # Buscar inventário de forma limpa e segura
    try:
        inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()
    except Exception:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE inventario ADD COLUMN IF NOT EXISTS data_limite TIMESTAMP;"))
            conn.commit()
        inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()

    # Itens esperados vs Registrados
    esperados = Patrimonio.query.filter_by(sala_id=inv.sala_id, status='ativo').all()
    registrados = ItemInventario.query.filter_by(inventario_id=inv.id).all()
    
    confirmados = [r for r in registrados if r.status == 'confirmado']
    
    # Proteção para p.id vs ItemInventario sem patrimônio id (tags desconhecidas)
    registrados_ids = [r.patrimonio_id for r in registrados if r.patrimonio_id is not None]
    faltantes = [p for p in esperados if p.id not in registrados_ids]
    
    fora_de_lugar = [r for r in registrados if r.status == 'fora_de_lugar']
    nao_encontrados = [r for r in registrados if r.status == 'nao_encontrado']
    
    return render_template('admin/balanco.html', 
                           inv=inv, 
                           confirmados=confirmados, 
                           faltantes=faltantes, 
                           fora_de_lugar=fora_de_lugar,
                           nao_encontrados=nao_encontrados)

@app.route('/admin/validar_inventario/<int:inv_id>')
@login_required
def admin_validar_inventario(inv_id):
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()
    inv.status = 'validado'
    db.session.commit()
    flash('Inventário validado e arquivado!', 'success')
    return redirect(url_for('admin_dashboard'))

# CRUD Professores (Com E-mail Real)
@app.route('/admin/responsaveis', methods=['GET', 'POST'])
@login_required
def admin_responsaveis():
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    e_id = current_user.escola_id
    if request.method == 'POST':
        senha_pura = request.form.get('password')
        novo_u = User(
            username=request.form.get('username'),
            email=request.form.get('email'),
            nome=request.form.get('nome'),
            role=request.form.get('role', 'professor'),
            escola_id=e_id
        )
        novo_u.set_password(senha_pura)
        
        sala_ids = request.form.getlist('sala_ids')
        for sid in sala_ids:
            s = Sala.query.filter_by(id=sid, escola_id=e_id).first()
            if s: novo_u.salas.append(s)
        
        db.session.add(novo_u)
        db.session.commit()
        
        # Enviar e-mail de boas-vindas com credenciais e ambientes
        base_url = request.host_url.rstrip('/')
        enviar_email(
            novo_u.email,
            f'Seu acesso ao SENAI Patrimônio — {current_user.escola.nome}',
            corpo_boas_vindas(novo_u, senha_pura, current_user.escola.nome, base_url)
        )
        
        flash(f'Conta criada e e-mail enviado para {novo_u.email}!', 'success')
        return redirect(url_for('admin_responsaveis'))
    
    responsaveis = User.query.filter(User.role.in_(['professor', 'coordenador']), User.escola_id == e_id).all()
    salas = Sala.query.filter_by(escola_id=e_id).all()
    return render_template('admin/responsaveis.html', responsaveis=responsaveis, salas=salas)

# Validação de Relocação
@app.route('/admin/relocacao/<int:id>/<string:acao>')
@login_required
def admin_processar_relocacao(id, acao):
    if current_user.role not in ['admin', 'coordenador']: return redirect(url_for('index'))
    req = SolicitacaoRealocacao.query.filter_by(id=id, escola_id=current_user.escola_id).first_or_404()
    if acao == 'aprovar':
        req.status = 'aprovado'
        req.patrimonio.sala_id = req.sala_destino_id
    else:
        req.status = 'recusado'
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

# --- Dashboard Professor ---
@app.route('/professor')
@login_required
def professor_dashboard():
    if current_user.role != 'professor':
        return redirect(url_for('index'))
    
    salas_data = []
    for s in current_user.salas:
        # Itens que DEVERIAM estar aqui
        total_ativos = Patrimonio.query.filter_by(sala_id=s.id, status='ativo').count()
        
        # Último inventário (para progresso e prazo)
        inv = Inventario.query.filter_by(sala_id=s.id).order_by(Inventario.data_hora_inicio.desc()).first()
        
        # Última contagem FINALIZADA
        last_inv = Inventario.query.filter_by(sala_id=s.id, status='concluido').order_by(Inventario.data_hora_fim.desc()).first()
        
        # Métricas do inventário atual
        encontrados = 0
        conflitos = 0
        quebrados = Patrimonio.query.filter_by(sala_id=s.id, status_conservacao='quebrado').count()
        
        if inv:
            # Encontrados: itens no inventário que pertencem a esta sala
            encontrados = db.session.query(func.count(ItemInventario.id)).\
                join(Patrimonio, ItemInventario.patrimonio_id == Patrimonio.id).\
                filter(ItemInventario.inventario_id == inv.id, Patrimonio.sala_id == s.id).scalar() or 0
                
            # Conflitos: itens achados aqui que são de outra sala
            conflitos = db.session.query(func.count(ItemInventario.id)).\
                join(Patrimonio, ItemInventario.patrimonio_id == Patrimonio.id).\
                filter(ItemInventario.inventario_id == inv.id, Patrimonio.sala_id != s.id).scalar() or 0

        acuracia = min(100, round((encontrados / total_ativos) * 100)) if total_ativos > 0 else 100
        
        salas_data.append({
            'obj': s,
            'acuracia': acuracia,
            'encontrados': encontrados,
            'faltantes': max(0, total_ativos - encontrados),
            'conflitos': conflitos,
            'quebrados': quebrados,
            'total_ativos': total_ativos,
            'ultima_contagem': last_inv.data_hora_fim.strftime('%d/%m/%Y') if last_inv and last_inv.data_hora_fim else 'N/A',
            'prazo': inv.data_limite.strftime('%d/%m/%Y') if inv and inv.data_limite else 'Sem prazo'
        })
        
    return render_template('responsavel/dashboard.html', salas_meta=salas_data)

# Inventário Logic
@app.route('/inventario/sala/<int:sala_id>')
@login_required
def inventario_sala(sala_id):
    sala = Sala.query.filter_by(id=sala_id, escola_id=current_user.escola_id).first_or_404()
    if current_user.role != 'admin' and sala not in current_user.salas:
        flash('Acesso negado!', 'danger')
        return redirect(url_for('index'))
    
    # Buscar inventário de forma limpa e segura
    try:
        inv = Inventario.query.filter_by(sala_id=sala_id, status='iniciado', escola_id=current_user.escola_id).first()
    except:
        # Recuperação silenciosa se houver falha de esquema
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text("ALTER TABLE inventario ADD COLUMN IF NOT EXISTS data_limite TIMESTAMP;"))
            conn.commit()
        inv = Inventario.query.filter_by(sala_id=sala_id, status='iniciado', escola_id=current_user.escola_id).first()

    if not inv:
        inv = Inventario(sala_id=sala_id, responsavel_id=current_user.id, escola_id=current_user.escola_id)
        db.session.add(inv)
        db.session.commit()
    
    itens_esperados = Patrimonio.query.filter_by(sala_id=sala_id, status='ativo').all()
    conferidos_ids = [item.patrimonio_id for item in ItemInventario.query.filter_by(inventario_id=inv.id).all()]
    
    itens_json = [{"id": p.id, "numero": str(p.numero_patrimonio), "descricao": str(p.descricao)} for p in itens_esperados]
    
    return render_template('inventario.html', 
                           sala=sala, 
                           inventario=inv, 
                           itens_esperados=itens_json, 
                           conferidos_ids=conferidos_ids)

# Registro de Ações de Balanço
@app.route('/inventario/item/foto/<int:item_id>', methods=['POST'])
@login_required
def inventario_item_foto(item_id):
    item = ItemInventario.query.get_or_404(item_id)
    # Segurança: Verificar se pertence à escola
    if item.inventario.escola_id != current_user.escola_id: return "Erro", 403
    
    foto_url = request.form.get('foto_url')
    if foto_url:
        item.foto_validacao_url = foto_url
        db.session.commit()
        return jsonify({"status": "ok"})
    return jsonify({"status": "erro"}), 400

@app.route('/inventario/assinar/<int:inv_id>', methods=['POST'])
@login_required
def inventario_assinar(inv_id):
    inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()
    assinatura = request.form.get('assinatura_base64')
    
    if assinatura:
        inv.assinatura_base64 = assinatura
        inv.status = 'concluido'
        inv.data_hora_fim = datetime.utcnow()
        db.session.commit()
        flash('Balanço assinado e enviado com sucesso!', 'success')
        return redirect(url_for('professor_dashboard'))
    
    flash('Assinatura é obrigatória para finalizar.', 'danger')
    return redirect(url_for('inventario_sala', sala_id=inv.sala_id))

@app.route('/inventario/subir_excel/<int:inv_id>', methods=['POST'])
@login_required
def inventario_subir_excel(inv_id):
    inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado', 'danger')
        return redirect(url_for('inventario_sala', sala_id=inv.sala_id))
        
    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        count = 0
        
        # O Excel exportado tem numero de patrimonio na primeira coluna (A)
        for row in sheet.iter_rows(min_row=2):
            numero = str(row[0].value).strip() if row[0].value else None
            if not numero: continue
            
            pat = Patrimonio.query.filter_by(numero_patrimonio=numero, escola_id=current_user.escola_id).first()
            if not pat: continue
            
            # Verifica se já foi conferido
            existente = ItemInventario.query.filter_by(inventario_id=inv_id, patrimonio_id=pat.id).first()
            if existente: continue
            
            if pat.sala_id == inv.sala_id:
                item_inv = ItemInventario(inventario_id=inv_id, patrimonio_id=pat.id, sala_id_da_vez=inv.sala_id, status='confirmado')
            else:
                # ITEM DE OUTRO AMBIENTE: Criar ItemInventario e abrir Solicitação de Realocação
                item_inv = ItemInventario(inventario_id=inv_id, patrimonio_id=pat.id, sala_id_da_vez=inv.sala_id, status='fora_de_lugar')
                # Gera solicitação para o Admin autorizar a mudança definitiva
                solicitacao = SolicitacaoRealocacao(
                    escola_id=inv.escola_id,
                    patrimonio_id=pat.id,
                    sala_origem_id=pat.sala_id,
                    sala_destino_id=inv.sala_id,
                    responsavel_id=current_user.id,
                    status='pendente',
                    observacao='Detectado via Balanço Excel'
                )
                db.session.add(solicitacao)
                
            db.session.add(item_inv)
            count += 1
            
        db.session.commit()
        flash(f'Excel processado! {count} itens sincronizados com sucesso.', 'success')
    except Exception as e:
        flash(f'Erro ao processar arquivo: {e}', 'danger')
        
    return redirect(url_for('inventario_sala', sala_id=inv.sala_id))

# Registro de Scan (AJAX) com verificação de local
@app.route('/inventario/scan/<int:inv_id>', methods=['POST'])
@login_required
def inventario_registrar_scan(inv_id):
    inv = Inventario.query.filter_by(id=inv_id, escola_id=current_user.escola_id).first_or_404()
    numero = request.form.get('numero')
    sala_id = inv.sala_id
    
    pat = Patrimonio.query.filter_by(numero_patrimonio=numero, escola_id=current_user.escola_id).first()
    
    if not pat:
        item_inv = ItemInventario(inventario_id=inv_id, status='nao_encontrado', observacao=f'Tag {numero} não cadastrada.')
        db.session.add(item_inv)
        db.session.commit()
        return jsonify({
            'status': 'wrong_room', 
            'pat_id': pat.id,
            'descricao': pat.descricao,
            'sala_atual': pat.sala.nome,
            'imagem': pat.imagem_url,
            'message': f'Item pertence à {pat.sala.nome}. Deseja solicitar relocação?'
        })

@app.route('/inventario/marcar_quebrado', methods=['POST'])
@login_required
def marcar_quebrado():
    data = request.json
    pat = Patrimonio.query.filter_by(id=data['pat_id'], escola_id=current_user.escola_id).first_or_404()
    pat.status_conservacao = 'quebrado'
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/inventario/solicitar_relocacao', methods=['POST'])
@login_required
def solicitar_relocacao():
    data = request.json
    nova_sol = SolicitacaoRealocacao(
        patrimonio_id=data['pat_id'],
        sala_origem_id=data['sala_origem_id'] if data.get('sala_origem_id') else Patrimonio.query.get(data['pat_id']).sala_id,
        sala_destino_id=data['sala_destino_id'],
        responsavel_id=current_user.id,
        escola_id=current_user.escola_id,
        observacao="Encontrado durante inventário."
    )
    item_inv = ItemInventario(
        inventario_id=data['inv_id'],
        patrimonio_id=data['pat_id'],
        sala_id_da_vez=data['sala_destino_id'],
        status='fora_de_lugar'
    )
    db.session.add(nova_sol)
    db.session.add(item_inv)
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/inventario/finalizar/<int:inv_id>')
@login_required
def finalizar_inventario(inv_id):
    inv = Inventario.query.get_or_404(inv_id)
    inv.status = 'concluido'
    inv.data_hora_fim = datetime.utcnow()
    db.session.commit()
    flash('Inventário finalizado!', 'success')
    return redirect(url_for('professor_dashboard'))

# --- Chat Logic ---
@app.route('/chat/<int:sala_id>', methods=['GET', 'POST'])
@login_required
def chat(sala_id):
    # Segurança: Sala deve pertencer à escola do usuário
    sala = Sala.query.filter_by(id=sala_id, escola_id=current_user.escola_id).first_or_404()
    
    if request.method == 'POST':
        msg = MensagemChat(
            escola_id=current_user.escola_id,
            sala_id=sala_id,
            usuario_id=current_user.id,
            usuario_tipo='adm' if current_user.role in ['admin', 'coordenador'] else 'professor',
            texto=request.form.get('texto')
        )
        db.session.add(msg)
        db.session.commit()
        return jsonify({"status": "ok"})
    
    # Marcar mensagens como lidas para este usuário
    MensagemChat.query.filter_by(sala_id=sala_id).filter(MensagemChat.usuario_id != current_user.id).update({MensagemChat.lida: True})
    db.session.commit()
    
    mensagens = MensagemChat.query.filter_by(sala_id=sala_id).order_by(MensagemChat.data_hora.asc()).all()
    return render_template('chat.html', sala=sala, mensagens=mensagens)

# API para Chat (Poll)
@app.route('/api/chat/<int:sala_id>')
@login_required
def api_chat(sala_id):
    # Segurança: Sala deve pertencer à escola do usuário
    sala = Sala.query.filter_by(id=sala_id, escola_id=current_user.escola_id).first_or_404()
    
    mensagens = MensagemChat.query.filter_by(sala_id=sala_id).order_by(MensagemChat.data_hora.asc()).all()
    res = []
    for m in mensagens:
        res.append({
            'usuario': m.usuario.nome,
            'foto': m.usuario.foto_url or f"https://api.dicebear.com/7.x/avataaars/svg?seed={m.usuario.username}",
            'tipo': m.usuario_tipo,
            'texto': m.texto,
            'data': m.data_hora.strftime('%H:%M'),
            'meu': m.usuario_id == current_user.id
        })
    return jsonify(res)

# --- Inicialização Multi-Escola SENAI SP ---
# Estrutura de escolas da rede SENAI
ESCOLAS_SENAI = [
    {'nome': 'SENAI Morvan Figueiredo', 'codigo': '1.03', 'cidade': 'São Paulo', 'admin_user': 'admin.morvan', 'admin_email': 'admin.morvan@sp.senai.br'},
    # Adicione novas escolas aqui conforme a rede expandir:
    # {'nome': 'SENAI Escola X', 'codigo': '1.XX', 'cidade': 'Cidade', 'admin_user': 'admin.escolax', 'admin_email': 'admin.escolax@sp.senai.br'},
]

def seed():
    """Inicializa o banco com as escolas da rede SENAI configuradas acima."""
    for escola_data in ESCOLAS_SENAI:
        # Verificar se a escola já existe (pelo código único)
        escola = Escola.query.filter_by(codigo_senai=escola_data['codigo']).first()
        if not escola:
            escola = Escola(
                nome=escola_data['nome'],
                codigo_senai=escola_data['codigo'],
                cidade=escola_data['cidade']
            )
            db.session.add(escola)
            db.session.commit()
            print(f"[SEED] Escola criada: {escola_data['nome']} ({escola_data['codigo']})")

        # Criar admin da escola se não existir
        admin_existente = User.query.filter_by(username=escola_data['admin_user']).first()
        if not admin_existente:
            admin = User(
                username=escola_data['admin_user'],
                email=escola_data['admin_email'],
                nome=f"Administrador — {escola_data['nome']}",
                role='admin',
                escola_id=escola.id
            )
            admin.set_password('Senai@2025')
            db.session.add(admin)
            db.session.commit()
            print(f"[SEED] Admin criado: {escola_data['admin_user']} | Senha: Senai@2025")
    
    # Admin Master Global (acesso irrestrito, para suporte técnico)
    if not User.query.filter_by(username='admin').first():
        primeira_escola = Escola.query.first()
        if primeira_escola:
            master = User(
                username='admin',
                email='admin@sp.senai.br',
                nome='Administrador Master — Suporte TI',
                role='admin',
                escola_id=primeira_escola.id
            )
            master.set_password('admin123')
            db.session.add(master)
            db.session.commit()
            print("[SEED] Admin Master criado: admin | Senha: admin123")

# Verificação e Reset Forçado para Limpeza
with app.app_context():
    # Migração Manual Silenciosa para garantir novos campos no Railway/Postgres
    # Deve rodar ANTES de qualquer query no DB
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # Postgres: ADD COLUMN IF NOT EXISTS
            # Usando blocos individuais para garantir que uma falha em um não pare o outro
            try:
                conn.execute(text("ALTER TABLE inventario ADD COLUMN IF NOT EXISTS data_limite TIMESTAMP;"))
                conn.commit()
            except Exception as e: print(f"Erro data_limite: {e}")
            
            try:
                conn.execute(text("ALTER TABLE mensagem_chat ADD COLUMN IF NOT EXISTS lida BOOLEAN DEFAULT FALSE;"))
                conn.commit()
            except Exception as e: print(f"Erro lida: {e}")
            
            print("Migração: Colunas verificadas/adicionadas com sucesso.")
    except Exception as e:
        print(f"Migração manual fatal: {e}")

    # Criar tabelas se não existirem
    db.create_all()
    
    # Migração de dados: atualizar escola genérica para SENAI Morvan Figueiredo
    try:
        escola_antiga = Escola.query.filter_by(codigo_senai='SP-01').first()
        if escola_antiga:
            escola_antiga.nome = 'SENAI Morvan Figueiredo'
            escola_antiga.codigo_senai = '1.03'
            escola_antiga.cidade = 'São Paulo'
            db.session.commit()
            print("[MIGRAÇÃO] Escola atualizada para: SENAI Morvan Figueiredo (1.03)")
    except Exception as e:
        print(f"[MIGRAÇÃO] Escola rename: {e}")
    
    seed()

if __name__ == '__main__':
    app.run(debug=True)
